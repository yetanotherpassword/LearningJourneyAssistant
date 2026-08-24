"""Unit tests for lja.data.excel_loader. Builds a small in-memory workbook
mirroring the real CSE_results_*.xlsx shape rather than depending on the
large real file, so these run fast and don't need a specific dataset
present on disk.
"""

from __future__ import annotations

import openpyxl
import pytest

from lja.data.excel_loader import _parse_silo_list, _split_assessment_type, load_dataset


def test_parse_silo_list_single_entry() -> None:
    assert _parse_silo_list("SILO3: reporting outcomes to an audience") == [
        ("SILO3", "reporting outcomes to an audience")
    ]


def test_parse_silo_list_multiple_entries_preserves_order() -> None:
    raw = "SILO2: abstract data types; SILO4: object-oriented design; SILO3: code reuse"
    assert _parse_silo_list(raw) == [
        ("SILO2", "abstract data types"),
        ("SILO4", "object-oriented design"),
        ("SILO3", "code reuse"),
    ]


@pytest.mark.parametrize(
    ("assessment_type", "expected"),
    [
        ("CSE1OOF - Test", ("CSE1OOF", "Test")),
        (
            "CSE3CAP - Assignment: Final written Project Report",
            ("CSE3CAP", "Assignment: Final written Project Report"),
        ),
    ],
)
def test_split_assessment_type(assessment_type: str, expected: tuple[str, str]) -> None:
    assert _split_assessment_type(assessment_type) == expected


@pytest.fixture
def workbook_path(tmp_path):
    """A minimal two-subject workbook with the real column headers, small
    enough to read in the test but exercising the same parsing paths as the
    real dataset -- including the "Average Total" column-name collision
    fixed in excel_loader.py.
    """
    wb = openpyxl.Workbook()

    am = wb.active
    am.title = "Assessment Map"
    am.append(["Subject Code", "Assessment Type", "Weight", "Contribution", "Early Assessment", "Hurdle", "SILOs", "SILO Theme Summary"])
    am.append(["SUBA", "SUBA - Test", 0.4, "Individual", "Yes", "No", "SILO1, SILO2", "SILO1: alpha skill; SILO2: beta skill"])
    am.append(["SUBB", "SUBB - Exam", 0.6, "Individual", "No", "No", "SILO1", "SILO1: gamma skill"])

    res = wb.create_sheet("Results")
    res.append(["Student ID", "Assessment Type", "Score (1-100)", "Feedback Comment", "SILO's", "Weight", "Weighted Score"])
    res.append(["STU0001", "SUBA - Test", 80, "Good work.", "SILO1: alpha skill; SILO2: beta skill", 0.4, 32.0])
    res.append(["STU0001", "SUBB - Exam", 40, "Needs work.", "SILO1: gamma skill", 0.6, 24.0])
    res.append(["STU0002", "SUBA - Test", 30, "Weak.", "SILO1: alpha skill; SILO2: beta skill", 0.4, 12.0])
    res.append(["STU0002", "SUBB - Exam", 35, "Weak.", "SILO1: gamma skill", 0.6, 21.0])

    ss = wb.create_sheet("Student Summary")
    ss.append(["Student ID", "SUBA Total", "SUBB Total", "Average Total", "Performance Band"])
    ss.append(["STU0001", 80.0, 40.0, 60.0, "C range"])
    ss.append(["STU0002", 30.0, 35.0, 32.5, "At risk"])

    path = tmp_path / "mini.xlsx"
    wb.save(path)
    return str(path)


def test_load_dataset_parses_silos_deduplicated(workbook_path: str) -> None:
    dataset = load_dataset(workbook_path)
    assert set(dataset.silos.keys()) == {"SUBA:SILO1", "SUBA:SILO2", "SUBB:SILO1"}
    assert dataset.silos["SUBA:SILO1"].text == "alpha skill"
    # Same local id, different subject -> different SILO, not merged.
    assert dataset.silos["SUBB:SILO1"].text == "gamma skill"


def test_load_dataset_assessments(workbook_path: str) -> None:
    dataset = load_dataset(workbook_path)
    assert len(dataset.assessments) == 2
    suba = next(a for a in dataset.assessments if a.subject_code == "SUBA")
    assert suba.assessment_name == "Test"
    assert suba.silo_ids == ("SILO1", "SILO2")
    assert suba.early_assessment is True
    assert suba.hurdle is False


def test_load_dataset_results(workbook_path: str) -> None:
    dataset = load_dataset(workbook_path)
    assert len(dataset.results) == 4
    row = next(r for r in dataset.results if r.student_id == "STU0001" and r.subject_code == "SUBA")
    assert row.assessment_name == "Test"
    assert row.score == 80.0
    assert row.silo_ids == ("SILO1", "SILO2")


def test_load_dataset_student_summary_excludes_average_total(workbook_path: str) -> None:
    """Regression test: 'Average Total' also ends with ' Total' and must not
    leak into subject_totals as a fake subject called 'Average'.
    """
    dataset = load_dataset(workbook_path)
    stu1 = next(s for s in dataset.student_summaries if s.student_id == "STU0001")
    assert stu1.subject_totals == {"SUBA": 80.0, "SUBB": 40.0}
    assert "Average" not in stu1.subject_totals
    assert stu1.average_total == 60.0
    assert stu1.performance_band == "C range"
